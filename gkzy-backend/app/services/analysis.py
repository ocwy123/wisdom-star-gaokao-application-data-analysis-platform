from flask import Blueprint, request, jsonify, current_app
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, and_, or_, desc, asc
from datetime import datetime
from functools import wraps
from app.extensions import db
import json
from app.models import School, Major, SchoolMajor, AdmRecord, SchoolHeat, MajorEmployment, ScoreSegment
import json
from datetime import datetime
analysis_bp = Blueprint('analysis', __name__, url_prefix='/api/analysis')

# ==================== 深度信息检索接口 ====================

@analysis_bp.route('/search', methods=['POST'])
def deep_search():
    """
    深度信息检索
    请求体：{
        "keyword": "搜索关键词",
        "types": ["school", "major", "employment"],  # 检索类型
        "filters": {  # 筛选条件
            "province": "北京市",
            "type": "理工类",
            "score_range": [500, 700],
            "year": 2024
        },
        "page": 1,
        "page_size": 20,
        "sort_by": "heat_score",
        "sort_order": "desc"
    }
    """
    try:
        data = request.get_json()
        keyword = data.get('keyword', '')
        search_types = data.get('types', ['school', 'major', 'employment'])
        filters = data.get('filters', {})
        page = data.get('page', 1)
        page_size = data.get('page_size', 20)
        sort_by = data.get('sort_by', 'relevance')
        sort_order = data.get('sort_order', 'desc')
        
        results = []
        
        # 1. 学校检索
        if 'school' in search_types:
            schools = search_schools(keyword, filters)
            results.extend(schools)
        
        # 2. 专业检索
        if 'major' in search_types:
            majors = search_majors(keyword, filters)
            results.extend(majors)
        
        # 3. 就业数据检索
        if 'employment' in search_types:
            employment = search_employment(keyword, filters)
            results.extend(employment)
        
        # 4. 招生记录检索
        if 'admission' in search_types:
            admissions = search_admissions(keyword, filters)
            results.extend(admissions)
        
        # 5. 热度数据检索
        if 'heat' in search_types:
            heat_data = search_heat_data(keyword, filters)
            results.extend(heat_data)
        
        # 排序
        results = sort_results(results, sort_by, sort_order)
        
        # 分页
        total = len(results)
        start = (page - 1) * page_size
        end = start + page_size
        paginated_results = results[start:end]
        
        return jsonify({
            'success': True,
            'data': {
                'items': paginated_results,
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size
            }
        })
        
    except Exception as e:
        print(f"搜索错误：{e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def search_schools(keyword, filters):
    """搜索学校"""
    from app.services.admin_auth import School
    from sqlalchemy import case
    
    query = School.query
    
    # 关键词搜索
    if keyword:
        query = query.filter(
            or_(
                School.name.contains(keyword),
                School.code.contains(keyword),
                School.city.contains(keyword)
            )
        )
        
        # 添加相关性排序：完全匹配 > 开头匹配 > 包含匹配
        # 同一优先级下，优先显示 985/211/双一流学校
        relevance_score = case(
            (School.name == keyword, 3),
            (School.name.like(keyword + '%'), 2),
            (School.name.like('%' + keyword + '%'), 1),
            else_=0
        )
        
        # 学校标签加分：985/211/双一流学校获得额外加分
        school_priority = case(
            (and_(School.is_985, School.is_211, School.is_double_first), 10),  # 三者都是
            (and_(School.is_985, School.is_211), 9),  # 985+211
            (and_(School.is_985, School.is_double_first), 9),  # 985+双一流
            (and_(School.is_211, School.is_double_first), 9),  # 211+双一流
            (School.is_985, 8),  # 仅 985
            (School.is_211, 7),  # 仅 211
            (School.is_double_first, 6),  # 仅双一流
            else_=0
        )
        
        # 先按相关性排序，再按学校优先级排序，最后按名称排序
        query = query.order_by(desc(relevance_score), desc(school_priority), School.name)
    
    # 筛选条件
    if filters.get('province'):
        query = query.filter(School.province == filters['province'])
    
    if filters.get('city'):
        query = query.filter(School.city.contains(filters['city']))
    
    if filters.get('school_type'):
        query = query.filter(School.type == filters['school_type'])
    
    if 'is_985' in filters:
        query = query.filter(School.is_985 == filters['is_985'])
    
    if 'is_211' in filters:
        query = query.filter(School.is_211 == filters['is_211'])
    
    schools = query.limit(50).all()
    
    results = []
    for school in schools:
        # 获取学校热度
        heat = get_school_heat(school.id)
        
        # 计算相关性（学校名称完全匹配获得最高分）
        relevance = calculate_relevance(keyword, school.name)
        
        # 如果是完全匹配的学校名称，给予额外加分
        if school.name == keyword:
            relevance = 1.0
        elif school.name.startswith(keyword):
            relevance = max(relevance, 0.8)
        
        results.append({
            'id': school.id,
            'type': 'school',
            'name': school.name,
            'school_name': school.name,
            'code': school.code,
            'province': school.province,
            'city': school.city,
            'type_name': school.type,
            'is_985': school.is_985,
            'is_211': school.is_211,
            'is_double_first': school.is_double_first,
            'description': school.description,
            'heat_score': heat.get('heat_score', 0),
            'search_count': heat.get('search_count', 0),
            'favorite_count': heat.get('favorite_count', 0),
            'relevance': relevance
        })
    
    return results

def search_majors(keyword, filters):
    """搜索专业"""
    from app.services.admin_auth import Major
    
    query = Major.query
    
    if keyword:
        query = query.filter(
            or_(
                Major.name.contains(keyword),
                Major.code.contains(keyword),
                Major.degree.contains(keyword)
            )
        )
    
    majors = query.limit(50).all()
    
    results = []
    for major in majors:
        # 获取专业就业数据
        employment = get_major_employment(major.id)
        
        # 计算相关性（专业数据的相关性应该低于学校数据）
        relevance = calculate_relevance(keyword, major.name)
        if major.name == keyword:
            relevance = 0.8  # 专业名称完全匹配，但低于学校类型的结果
        elif major.name.startswith(keyword):
            relevance = max(relevance, 0.7)  # 专业名称开头匹配
        
        results.append({
            'id': major.id,
            'type': 'major',
            'name': major.name,
            'major_name': major.name,
            'code': major.code,
            'duration': major.duration,
            'degree': major.degree,
            'subjects': major.subjects,
            'description': major.description,
            'avg_salary': employment.get('avg_salary', 0),
            'employment_rate': employment.get('employment_rate', 0),
            'relevance': relevance
        })
    
    return results

def search_employment(keyword, filters):
    """搜索就业数据"""
    from app.services.admin_auth import MajorEmployment, Major
    
    query = MajorEmployment.query.join(
        Major, MajorEmployment.major_id == Major.id
    )
    
    if keyword:
        query = query.filter(Major.name.contains(keyword))
    
    if filters.get('year'):
        query = query.filter(MajorEmployment.year == filters['year'])
    
    if filters.get('min_salary'):
        query = query.filter(MajorEmployment.avg_salary >= filters['min_salary'])
    
    employment_data = query.limit(50).all()
    
    results = []
    for emp in employment_data:
        major = Major.query.get(emp.major_id)
        
        # 计算相关性（就业数据的相关性应该低于学校数据）
        relevance = 0.0
        if major:
            if major.name == keyword:
                relevance = 0.7  # 专业名称完全匹配，但低于学校类型的结果
            elif major.name.startswith(keyword):
                relevance = 0.6  # 专业名称开头匹配
            elif keyword in major.name:
                relevance = 0.5  # 专业名称包含关键词
        
        results.append({
            'id': emp.id,
            'type': 'employment',
            'major_id': emp.major_id,
            'major_name': major.name if major else '未知',
            'year': emp.year,
            'avg_salary': emp.avg_salary,
            'industry_distribution': json.loads(emp.industry_distribution) if emp.industry_distribution else {},
            'post_distribution': json.loads(emp.post_distribution) if emp.post_distribution else {},
            'region_distribution': json.loads(emp.region_distribution) if emp.region_distribution else {},
            'prospect': emp.prospect,
            'relevance': relevance
        })
    
    return results

def search_admissions(keyword, filters):
    """搜索招生记录"""
    from app.services.admin_auth import AdmRecord, School, Major
    
    query = AdmRecord.query.join(
        School, AdmRecord.school_id == School.id
    )
    
    if keyword:
        query = query.filter(
            or_(
                School.name.contains(keyword),
                AdmRecord.major_name.contains(keyword),
                AdmRecord.province.contains(keyword)
            )
        )
    
    if filters.get('province'):
        query = query.filter(AdmRecord.province == filters['province'])
    
    if filters.get('year'):
        query = query.filter(AdmRecord.year == filters['year'])
    
    if filters.get('batch'):
        query = query.filter(AdmRecord.batch == filters['batch'])
    
    if filters.get('subject'):
        query = query.filter(AdmRecord.subject == filters['subject'])
    
    if filters.get('score_range'):
        min_score, max_score = filters['score_range']
        query = query.filter(
            AdmRecord.min_score.between(min_score, max_score)
        )
    
    admissions = query.limit(50).all()
    
    results = []
    for adm in admissions:
        school = School.query.get(adm.school_id)
        
        # 计算相关性（招生数据的相关性应该低于学校数据）
        relevance = 0.0
        if school and school.name == keyword:
            relevance = 0.7  # 学校名称完全匹配，但低于学校类型的结果
        elif school and school.name.startswith(keyword):
            relevance = 0.6  # 学校名称开头匹配，但低于学校类型的结果
        elif school and keyword in school.name:
            relevance = 0.5  # 学校名称包含关键词
        elif keyword in adm.major_name:
            relevance = 0.4  # 专业名称包含关键词
        elif keyword in adm.province:
            relevance = 0.3  # 省份包含关键词
        
        results.append({
            'id': adm.id,
            'type': 'admission',
            'school_id': adm.school_id,
            'school_name': school.name if school else '未知',
            'major_name': adm.major_name,
            'major_second_name': adm.major_second_name,
            'province': adm.province,
            'year': adm.year,
            'plan_count': adm.plan_count,
            'subject': adm.subject,
            'batch': adm.batch,
            'major_group': adm.major_group,
            'min_score': adm.min_score,
            'min_rank': adm.min_rank,
            'relevance': relevance
        })
    
    return results

def search_heat_data(keyword, filters):
    """搜索热度数据"""
    from app.services.admin_auth import SchoolHeat, School
    
    query = SchoolHeat.query.join(
        School, SchoolHeat.school_id == School.id
    )
    
    if keyword:
        query = query.filter(School.name.contains(keyword))
    
    heat_data = query.limit(50).all()
    
    results = []
    for heat in heat_data:
        school = School.query.get(heat.school_id)
        
        # 计算相关性（热度数据的相关性应该低于学校数据）
        relevance = 0.0
        if school:
            if school.name == keyword:
                relevance = 0.6  # 学校名称完全匹配，但低于学校类型的结果
            elif school.name.startswith(keyword):
                relevance = 0.5  # 学校名称开头匹配
            elif keyword in school.name:
                relevance = 0.4  # 学校名称包含关键词
        
        results.append({
            'id': heat.id,
            'type': 'heat',
            'school_id': heat.school_id,
            'school_name': school.name if school else '未知',
            'search_count': heat.search_count,
            'favorite_count': heat.favorite_count,
            'view_count': heat.view_count,
            'heat_score': float(heat.heat_score) if heat.heat_score else 0,
            'relevance': relevance
        })
    
    return results

def get_school_heat(school_id):
    """获取学校热度"""
    from app.services.admin_auth import SchoolHeat
    
    heat = SchoolHeat.query.filter_by(school_id=school_id).first()
    if heat:
        return {
            'heat_score': float(heat.heat_score) if heat.heat_score else 0,
            'search_count': heat.search_count,
            'favorite_count': heat.favorite_count,
            'view_count': heat.view_count
        }
    return {}

def get_major_employment(major_id):
    """获取专业就业数据"""
    from app.services.admin_auth import MajorEmployment
    
    emp = MajorEmployment.query.filter_by(major_id=major_id).first()
    if emp:
        return {
            'avg_salary': emp.avg_salary,
            'employment_rate': 0  # 需要根据实际数据计算
        }
    return {}

def calculate_relevance(keyword, text):
    """计算相关性得分"""
    if not keyword or not text:
        return 0
    keyword = keyword.lower()
    text = text.lower()
    
    # 完全匹配：最高优先级
    if text == keyword:
        return 1.0
    
    # 开头匹配：高优先级
    if text.startswith(keyword):
        return 0.9
    
    # 包含关键词：根据位置计算相关性
    if keyword in text:
        position = text.index(keyword)
        return 0.8 - (position / len(text))
    
    return 0

def sort_results(results, sort_by, sort_order):
    """排序结果"""
    reverse = (sort_order == 'desc')
    
    if sort_by == 'relevance':
        # 按相关性排序，当 relevance 相同时，优先显示 985/211/双一流学校
        def get_sort_key(item):
            relevance = item.get('relevance', 0)
            
            # 计算学校标签优先级
            school_priority = 0
            if item.get('type') == 'school':
                if item.get('is_985') and item.get('is_211') and item.get('is_double_first'):
                    school_priority = 10
                elif item.get('is_985') and item.get('is_211'):
                    school_priority = 9
                elif item.get('is_985'):
                    school_priority = 8
                elif item.get('is_211'):
                    school_priority = 7
                elif item.get('is_double_first'):
                    school_priority = 6
            
            # 返回元组：(相关性，学校优先级)
            return (relevance, school_priority)
        
        return sorted(results, key=get_sort_key, reverse=reverse)
    elif sort_by == 'heat_score':
        return sorted(results, key=lambda x: x.get('heat_score', 0), reverse=reverse)
    elif sort_by == 'avg_salary':
        return sorted(results, key=lambda x: x.get('avg_salary', 0), reverse=reverse)
    elif sort_by == 'min_score':
        return sorted(results, key=lambda x: x.get('min_score', 0), reverse=reverse)
    elif sort_by == 'name':
        return sorted(results, key=lambda x: x.get('name', ''), reverse=reverse)
    
    return results

# ==================== 多维对比分析接口 ====================

@analysis_bp.route('/compare', methods=['POST'])
def multi_dimension_compare():
    """
    多维对比分析
    请求体：{
        "dimension": "school",  # 对比维度（单选）
        "metrics": ["avg_score", "heat_score", "admission_rate"],  # 指标列表
        "filters": {  # 筛选条件
            "school_ids": [1, 2, 3],
            "major_ids": [1, 2],
            "provinces": ["北京市", "上海市"]
        },
        "time_range": [2020, 2024]
    }
    """
    try:
        data = request.get_json()
        dimension = data.get('dimension', 'school')
        metrics = data.get('metrics', ['avg_score', 'heat_score'])
        filters = data.get('filters', {})
        time_range = data.get('time_range', [2020, 2024])
        
        result = {
            'dimension': dimension,
            'metrics': metrics,
            'data': []
        }
        
        # 根据维度调用不同的分析函数
        if dimension == 'school':
            result['data'] = compare_schools(filters, metrics, time_range)
        elif dimension == 'major':
            result['data'] = compare_majors(filters, metrics, time_range)
        elif dimension == 'province':
            result['data'] = compare_by_province(filters, metrics, time_range)
        elif dimension == 'heat':
            result['data'] = analyze_heat_trend(filters, metrics, time_range)
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        print(f"对比分析错误：{e}")
        return jsonify({'success': False, 'message': str(e)}), 500
def compare_schools(filters, metrics, time_range):
    """学校对比分析"""
    from app.services.admin_auth import School, AdmRecord, SchoolHeat
    
    school_ids = filters.get('school_ids', [])
    if not school_ids:
        return []
    
    result = []
    for school_id in school_ids:
        school = School.query.get(school_id)
        if not school:
            continue
        
        item = {
            'dimension': 'school',
            'dimension_value': school.name,
            'school_id': school_id,
            'data': {}
        }
        
        # 获取招生数据
        admissions = AdmRecord.query.filter(
            AdmRecord.school_id == school_id,
            AdmRecord.year.between(time_range[0], time_range[1])
        ).all()
        
        if 'avg_score' in metrics:
            scores = [a.min_score for a in admissions if a.min_score]
            item['data']['avg_score'] = sum(scores) / len(scores) if scores else 0
        
        if 'admission_rate' in metrics:
            # 计算录取率（需要实际数据）
            item['data']['admission_rate'] = 0.75  # 示例值
        
        if 'min_score' in metrics:
            item['data']['min_score'] = min([a.min_score for a in admissions if a.min_score]) if admissions else 0
        
        if 'max_score' in metrics:
            item['data']['max_score'] = max([a.min_score for a in admissions if a.min_score]) if admissions else 0
        
        # 获取热度数据
        heat = SchoolHeat.query.filter_by(school_id=school_id).first()
        if heat:
            if 'heat_score' in metrics:
                item['data']['heat_score'] = float(heat.heat_score) if heat.heat_score else 0
            if 'search_count' in metrics:
                item['data']['search_count'] = heat.search_count or 0
            if 'favorite_count' in metrics:
                item['data']['favorite_count'] = heat.favorite_count or 0
            if 'view_count' in metrics:
                item['data']['view_count'] = heat.view_count or 0
        
        # 获取学校基本信息
        if 'phd_count' in metrics:
            item['data']['phd_count'] = school.phd_count or 0
        if 'master_count' in metrics:
            item['data']['master_count'] = school.master_count or 0
        if 'founded_year' in metrics:
            item['data']['founded_year'] = school.founded_year or 0
        if 'is_985' in metrics:
            item['data']['is_985'] = 1 if school.is_985 else 0
        if 'is_211' in metrics:
            item['data']['is_211'] = 1 if school.is_211 else 0
        if 'is_double_first' in metrics:
            item['data']['is_double_first'] = 1 if school.is_double_first else 0
        
        # 获取招生统计数据
        if 'admission_count' in metrics:
            item['data']['admission_count'] = len(admissions)
        if 'plan_count' in metrics:
            total_plan = sum([a.plan_count for a in admissions if a.plan_count])
            item['data']['plan_count'] = total_plan
        
        result.append(item)
    
    return result

def compare_majors(filters, metrics, time_range):
    """专业对比分析"""
    from app.services.admin_auth import Major, MajorEmployment, AdmRecord
    
    major_ids = filters.get('major_ids', [])
    if not major_ids:
        return []
    
    result = []
    for major_id in major_ids:
        major = Major.query.get(major_id)
        if not major:
            continue
        
        item = {
            'dimension': 'major',
            'dimension_value': major.name,
            'major_id': major_id,
            'data': {}
        }
        
        # 获取就业数据
        employment = MajorEmployment.query.filter_by(major_id=major_id).first()
        if employment:
            if 'avg_salary' in metrics:
                item['data']['avg_salary'] = employment.avg_salary
            if 'employment_rate' in metrics:
                item['data']['employment_rate'] = 0  # 需要实际数据
        
        # 获取招生数据（通过专业名称匹配）
        admissions = AdmRecord.query.filter(
            AdmRecord.major_name == major.name,
            AdmRecord.year.between(time_range[0], time_range[1])
        ).all()
        
        if 'avg_score' in metrics:
            scores = [a.min_score for a in admissions if a.min_score]
            item['data']['avg_score'] = sum(scores) / len(scores) if scores else 0
        
        result.append(item)
    
    return result

def compare_by_province(filters, metrics, time_range):
    """按省份对比分析"""
    from app.services.admin_auth import School, AdmRecord
    
    provinces = filters.get('provinces', [])
    if not provinces:
        # 获取所有省份
        provinces = db.session.query(School.province).distinct().all()
        provinces = [p[0] for p in provinces if p[0]]
    
    result = []
    for province in provinces[:10]:  # 限制数量
        schools = School.query.filter_by(province=province).all()
        school_ids = [s.id for s in schools]
        
        if not school_ids:
            continue
        
        admissions = AdmRecord.query.filter(
            AdmRecord.school_id.in_(school_ids),
            AdmRecord.year.between(time_range[0], time_range[1])
        ).all()
        
        item = {
            'dimension': 'province',
            'dimension_value': province,
            'data': {}
        }
        
        # 获取招生数据统计
        if 'avg_score' in metrics:
            scores = [a.min_score for a in admissions if a.min_score]
            item['data']['avg_score'] = sum(scores) / len(scores) if scores else 0
        
        if 'min_score' in metrics:
            scores = [a.min_score for a in admissions if a.min_score]
            item['data']['min_score'] = min(scores) if scores else 0
        
        if 'max_score' in metrics:
            scores = [a.min_score for a in admissions if a.min_score]
            item['data']['max_score'] = max(scores) if scores else 0
        
        if 'school_count' in metrics:
            item['data']['school_count'] = len(schools)
        
        if 'admission_count' in metrics:
            item['data']['admission_count'] = len(admissions)
        
        if 'plan_count' in metrics:
            total_plan = sum([a.plan_count for a in admissions if a.plan_count])
            item['data']['plan_count'] = total_plan
        
        # 获取省份学校统计信息
        if '985_count' in metrics:
            count_985 = sum([1 for s in schools if s.is_985])
            item['data']['985_count'] = count_985
        
        if '211_count' in metrics:
            count_211 = sum([1 for s in schools if s.is_211])
            item['data']['211_count'] = count_211
        
        if 'double_first_count' in metrics:
            count_double_first = sum([1 for s in schools if s.is_double_first])
            item['data']['double_first_count'] = count_double_first
        
        if 'city_count' in metrics:
            cities = set([s.city for s in schools if s.city])
            item['data']['city_count'] = len(cities)
        
        result.append(item)
    
    return result

def analyze_heat_trend(filters, metrics, time_range):
    """热度趋势分析"""
    from app.services.admin_auth import SchoolHeat, School
    
    query = SchoolHeat.query.join(
        School, SchoolHeat.school_id == School.id
    )
    
    if filters.get('province'):
        query = query.filter(School.province == filters['province'])
    
    if filters.get('school_type'):
        query = query.filter(School.type == filters['school_type'])
    
    heat_data = query.all()
    
    # 按热度分数排序
    heat_data.sort(key=lambda x: x.heat_score or 0, reverse=True)
    
    result = []
    for heat in heat_data[:20]:  # 前 20 名
        school = School.query.get(heat.school_id)
        
        item = {
            'dimension': 'heat',
            'dimension_value': school.name if school else '未知',
            'data': {
                'heat_score': float(heat.heat_score) if heat.heat_score else 0,
                'search_count': heat.search_count,
                'favorite_count': heat.favorite_count,
                'view_count': heat.view_count
            }
        }
        
        result.append(item)
    
    return result

# ==================== 高级筛选接口 ====================

@analysis_bp.route('/filters', methods=['GET'])
def get_filter_options():
    """获取筛选选项"""
    try:
        from app.services.admin_auth import School, Major
        
        # 获取省份列表
        provinces = db.session.query(School.province).distinct().all()
        provinces = [p[0] for p in provinces if p[0]]
        
        # 获取学校类型列表
        school_types = db.session.query(School.type).distinct().all()
        school_types = [t[0] for t in school_types if t[0]]
        
        # 获取专业列表
        majors = Major.query.all()
        major_list = [{'id': m.id, 'name': m.name} for m in majors]
        # 获取学校列表
        schools = School.query.all()
        school_list = [{'id': s.id, 'name': s.name} for s in schools]
        # 获取年份范围
        from app.services.admin_auth import AdmRecord
        years = db.session.query(AdmRecord.year).distinct().order_by(AdmRecord.year).all()
        years = [y[0] for y in years if y[0]]
        
        return jsonify({
            'success': True,
            'data': {
                'provinces': provinces,
                'school_types': school_types,
                'majors': major_list,
                'schools': school_list,
                'years': years,
                'score_ranges': [
                    {'min': 0, 'max': 400, 'label': '400 分以下'},
                    {'min': 400, 'max': 450, 'label': '400-450 分'},
                    {'min': 450, 'max': 500, 'label': '450-500 分'},
                    {'min': 500, 'max': 550, 'label': '500-550 分'},
                    {'min': 550, 'max': 600, 'label': '550-600 分'},
                    {'min': 600, 'max': 750, 'label': '600 分以上'}
                ]
            }
        })
        
    except Exception as e:
        print(f"获取筛选选项错误：{e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== 导出分析结果 ====================

@analysis_bp.route('/export', methods=['POST'])
def export_analysis():
    """导出分析结果"""
    try:
        import base64
        from io import BytesIO
        import pandas as pd
        
        data = request.get_json()
        export_type = data.get('type', 'excel')
        result_data = data.get('data', [])
        dimension = data.get('dimension', 'school')
        metrics = data.get('metrics', [])
        
        if not result_data:
            return jsonify({'success': False, 'message': '没有可导出的数据'})
        
        if export_type == 'json':
            return jsonify({
                'success': True,
                'data': result_data
            })
        
        elif export_type == 'excel':
            output = BytesIO()
            
            flat_data = []
            for item in result_data:
                row = {
                    '对比项': item.get('dimension_value', ''),
                    '维度类型': item.get('dimension', '')
                }
                
                item_data = item.get('data', {})
                for metric in metrics:
                    metric_value = item_data.get(metric, '')
                    row[get_metric_name(metric)] = metric_value
                
                flat_data.append(row)
            
            df = pd.DataFrame(flat_data)
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='分析结果')
                
                workbook = writer.book
                worksheet = writer.sheets['分析结果']
                
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                data_alignment = Alignment(horizontal="left", vertical="center")
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = data_alignment
                
                worksheet.insert_rows(1)
                title_cell = worksheet.cell(row=1, column=1, value=f'多维对比分析报告 - {get_dimension_name(dimension)}')
                from openpyxl.styles import Font as XLFont, Alignment as XLAlignment
                title_cell.font = XLFont(bold=True, size=16, color="1F4E79")
                title_cell.alignment = XLAlignment(horizontal="center", vertical="center")
                worksheet.merge_cells(f'A1:{chr(65 + len(df.columns) - 1)}1')
            
            output.seek(0)
            excel_data = output.getvalue()
            
            excel_base64 = base64.b64encode(excel_data).decode('utf-8')
            
            return jsonify({
                'success': True,
                'data': excel_base64,
                'format': 'xlsx',
                'filename': f'多维对比分析报告_{dimension}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            })
        
        elif export_type == 'csv':
            import csv
            from io import StringIO
            
            output = StringIO()
            if result_data:
                flat_data = []
                for item in result_data:
                    row = {
                        '对比项': item.get('dimension_value', ''),
                        '维度类型': item.get('dimension', '')
                    }
                    item_data = item.get('data', {})
                    for metric in metrics:
                        row[get_metric_name(metric)] = item_data.get(metric, '')
                    flat_data.append(row)
                
                writer = csv.DictWriter(output, fieldnames=['对比项', '维度类型'] + [get_metric_name(m) for m in metrics])
                writer.writeheader()
                writer.writerows(flat_data)
            
            return jsonify({
                'success': True,
                'data': output.getvalue(),
                'format': 'csv'
            })
        
        return jsonify({'success': False, 'message': '不支持的导出格式'})
        
    except Exception as e:
        print(f"导出错误：{e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'导出失败：{str(e)}'}), 500


def get_metric_name(metric_code):
    """获取指标的中文名称"""
    metric_names = {
        'avg_score': '平均分数',
        'heat_score': '热度分数',
        'admission_rate': '录取率',
        'min_score': '最低分数',
        'max_score': '最高分数',
        'search_count': '搜索次数',
        'favorite_count': '收藏次数',
        'view_count': '浏览次数',
        'phd_count': '博士点数量',
        'master_count': '硕士点数量',
        'founded_year': '建校年份',
        'is_985': '是否 985',
        'is_211': '是否 211',
        'is_double_first': '是否双一流',
        'admission_count': '招生数量',
        'plan_count': '计划人数',
        'avg_salary': '平均薪资',
        'employment_rate': '就业率',
        'school_count': '学校数量',
        '985_count': '985 院校数',
        '211_count': '211 院校数',
        'double_first_count': '双一流院校数',
        'city_count': '城市数量',
        'major_count': '专业数量',
        'province_count': '省份数量',
        'count': '数量'
    }
    return metric_names.get(metric_code, metric_code)


def get_dimension_name(dimension_code):
    """获取维度的中文名称"""
    dimension_names = {
        'school': '学校对比',
        'major': '专业对比',
        'province': '省份对比',
        'heat': '热度趋势'
    }
    return dimension_names.get(dimension_code, dimension_code)
